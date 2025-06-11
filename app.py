from flask import Flask, render_template, request, redirect, url_for, flash ,send_file
from datetime import datetime, timedelta
import os, re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from io import BytesIO
import pandas as pd
import pythoncom
import win32com.client as win32
import time
import psutil
app = Flask(__name__)
app.secret_key = 'your_secret_key'


# 주차 생성 함수 (2024년부터 현재까지)


def generate_weeks_from_2024():
    weeks = []
    start_date = datetime(2024, 1, 1)
    
    # 첫 월요일을 찾기
    while start_date.weekday() != 0:  # 0은 월요일
        start_date += timedelta(days=1)
    
    today = datetime.today()
    current_month = start_date.month  # 초기 월 설정
    week_count_in_month = 1  # 각 월의 첫 주차로 초기화

    # 2024년부터 현재 날짜까지의 주차 생성
    current_week = start_date
    while current_week + timedelta(days=7) <= today:
        year = current_week.year
        month = current_week.month

        # 새로운 달이면 주차를 1로 초기화
        if month != current_month:
            current_month = month
            week_count_in_month = 1  # 매월 첫 주차로 초기화

        # 주차 정보 생성
        week_str = f"{year}년 {month}월 {week_count_in_month}주차"
        
        # 52주를 초과하면 가장 오래된 주차 제거
        if len(weeks) >= 52:
            weeks.pop(0)
        
        # 주차 리스트에 추가
        weeks.append(week_str)
        
        # 다음 주차로 이동
        current_week += timedelta(days=7)
        week_count_in_month += 1  # 현재 월의 주차 증가

    # 현재 주차와 +1주차 생성
    for _ in range(2):  # 현재 주차와 +1주차까지 처리
        year = current_week.year
        month = current_week.month
        
        # 새로운 달이면 주차를 1로 초기화
        if month != current_month:
            current_month = month
            week_count_in_month = 1  # 매월 첫 주차로 초기화
        
        # 주차 정보 생성
        week_str = f"{year}년 {month}월 {week_count_in_month}주차"
        
        # 52주를 초과하면 가장 오래된 주차 제거
        if len(weeks) >= 52:
            weeks.pop(0)
        
        # 주차 리스트에 추가
        weeks.append(week_str)
        
        # 다음 주차로 이동
        current_week += timedelta(days=7)
        week_count_in_month += 1  # 현재 월의 주차 증가

    return weeks  # 전체 주차 리스트 반환

def generate_months():
    start_year = 2024
    current_year = datetime.now().year
    current_month = datetime.now().month
    months = []

    # 2024년부터 현재 월까지 월 추가
    for year in range(start_year, current_year + 1):
        for month in range(1, 13):
            if year == current_year and month > current_month:
                break
            months.append(f"{year}년 {month}월")  # 2자리로 월 표시

    # 최대 24개월로 제한
    if len(months) > 24:
        months = months[-24:]

    return months

def parse_week(week_string):
    # "2024년 1월 1주차" 같은 문자열을 year, month, week로 나눕니다.
    pattern = r"(\d{4})년 (\d{1,2})월 (\d)주차"
    match = re.match(pattern, week_string)
    
    if match:
        year = match.group(1)  # "2024"
        month = match.group(2)  # "1"
        week = match.group(3)  # "1"
        return year, month, week
    return None, None, None

from datetime import datetime, timedelta

def get_current_and_next_week():
    today = datetime.today()
    
    # 첫 월요일 찾기
    start_date = datetime(2024, 1, 1)
    while start_date.weekday() != 0:  # 월요일이 될 때까지 반복
        start_date += timedelta(days=1)

    # 현재 날짜 기준 주차 계산
    current_week = start_date
    current_month = current_week.month
    week_count_in_month = 1  # 각 월의 첫 주차로 초기화

    # 오늘 날짜까지 주차 이동
    while current_week + timedelta(days=7) <= today:
        if current_week.month != current_month:
            current_month = current_week.month
            week_count_in_month = 1  # 새로운 달의 첫 주차로 초기화
        else:
            week_count_in_month += 1  # 같은 달에서는 주차 증가

        current_week += timedelta(days=7)

    # 현재 주차 정보 생성
    current_year = current_week.year
    current_month = current_week.month
    current_week_str = f"{current_year}년 {current_month}월 {week_count_in_month}주차"

    # +1 주차 정보 생성
    next_week = current_week + timedelta(days=7)
    if next_week.month != current_month:
        week_count_in_month = 1  # 새로운 달의 첫 주차로 초기화
    else:
        week_count_in_month += 1

    next_year = next_week.year
    next_month = next_week.month
    next_week_str = f"{next_year}년 {next_month}월 {week_count_in_month}주차"

    return next_week_str


# 기본 페이지: 보고서 제출 폼
@app.route('/', methods=['GET', 'POST'])
def report_form():
    weeks = generate_weeks_from_2024()  # 2024년부터 현재까지 주차 목록 생성
    departments = ["혁신기획팀", "산업안전팀", "안전감사실", "경영지원팀","산장관광지","가평썰매장","칼봉산자연휴양림","연인산다목적캠핑장","자라섬캠핑장","교통약자이동지원센터","문화예술관","여성비전센터","종량제봉투판매","창업경제타운","한석봉체육관","생활체육파트","조종국민체육센터","청평호반문화체육센터","가평파크골프장"]   # 부서 목록
    current_week = get_current_and_next_week()
    selected_week = None
    selected_department = None
    months=generate_months()

    if request.method == 'POST':
        selected_week = request.form.get('week')
        selected_department = request.form.get('department')  # 사용자가 선택한 부서 값
        months=generate_months()

        # 폼 데이터를 처리 (파일 저장 등)
        flash(f'{selected_department} 부서의 보고서가 저장되었습니다.')

        # 폼 제출 후에도 선택한 값 유지
        return render_template('report_form.html', weeks=weeks, departments=departments, selected_week=selected_week, selected_department=selected_department,current_week=current_week,months=months)

    # GET 요청일 경우, 기본 폼 렌더링
    return render_template('report_form.html', weeks=weeks, departments=departments, selected_week=current_week, selected_department=selected_department,current_week=current_week,months=months)


# 보고서를 제출하면 데이터를 처리하는 경로
@app.route('/submit-report', methods=['POST'])
def submit_report():
    # 폼에서 제출된 데이터 가져오기
    department = request.form['department']
    task_name = request.form['task-name']
    target = request.form['target']
    date_start = request.form['date-start']
    date_end = request.form.get('date-end')
    location = request.form['location']
    method = request.form['method']
    task = request.form['task']
    work_type = request.form.get('workType')
    objective = request.form['objective']
    etc = request.form['etc']
    week = request.form['week']
    task_type = request.form.get('taskType')
    budget = request.form.get('budget')
    role = request.form.get('role')
    final_selection = request.form.get('finalSelection')
    docu = request.form.get('docu')
    process = request.form.get('additional-select-1')
    target_type = request.form.get('additional-select-2')
    months=generate_months()
    

    if target_type:
     final_target = f"{target}({target_type})"
    else:
      final_target = target

    # 주차 값 파싱
    year, month, week = parse_week(week)

    

    # date_start와 date_end를 결합해서 "기간" 생성
    if date_end and date_end != date_start:
        date_range = f"{date_start} ~ {date_end}"
    else:
        date_range = date_start

    # 추진내용 필터링
    data = []
    if target:
        data.append(('대상', target))
    if date_range:
        data.append(('기간', date_range))
    if location:
        data.append(('장소', location))
    if method:
        data.append(('기대효과/추진성과', method))
    if objective:
        data.append(('목적', objective))        
    if task:
        data.append(('내용', task))
    if etc:
        data.append(('행정사항', etc))
    if budget:
        data.append(('소요예산', budget))
    if docu :
        data.append(('관련자료',docu))    

    # 데이터 저장 형식 준비
    rows = [] 
    for label, value in data:
     if role:  # role 값이 있으면
        rows.append([f"{department}\n[{role}]", task_name, label, value])
     else:  # role 값이 없으면
        rows.append([f"{department}", task_name, label, value])

    # 파일 이름 설정
    filename = f"{target_type}_{process}_{final_selection}_{work_type}_{task_type}_{year}년_{month}월_{week}주차_{department}_report.xlsx"

    # 파일 처리
    if not os.path.exists(filename):
      wb = Workbook()
      ws = wb.active
      start_row = ws.max_row
    else:
      wb = load_workbook(filename)
      ws = wb.active
      start_row = ws.max_row + 1
   
    end_row = start_row + len(data) - 1

    # 데이터 추가
    for i, row in enumerate(rows):
        ws.append(row)

    # 행 병합 (Department와 Task Name 병합)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)

    # 테두리 스타일
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="thick", color="000000")

    # 테두리 및 정렬
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=4):
        for cell in row:
            top = thin
            bottom = thin
            left = thin
            right = thin

            if cell.row == start_row:
                top = thick
            if cell.row == end_row:
                bottom = thick

            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = None

    column_widths = {'A': 20, 'B': 30, 'C': 10, 'D': 50}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 파일 저장
    wb.save(filename)
    print(f"cc : {process}")
    flash(f'{department} 부서의 {year}년 {month}월 {week}주차 보고서가 성공적으로 제출되었습니다')

    # 제출 후 선택 값 유지하며 페이지를 렌더링
    weeks = generate_weeks_from_2024()
    departments = ["혁신기획팀", "산업안전팀", "안전감사실", "경영지원팀","산장관광지","가평썰매장","칼봉산자연휴양림","연인산다목적캠핑장","자라섬캠핑장","교통약자이동지원센터","문화예술관","여성비전센터","종량제봉투판매","창업경제타운","한석봉체육관","생활체육파트","조종국민체육센터","청평호반문화체육센터","가평파크골프장"]  # 부서 목록
    
    return render_template('report_form.html', weeks=weeks, departments=departments, selected_week=f"{year}년 {month}월 {week}주차", selected_department=department,months=months)


########################################### 데이터 취합 ############################
departments = ["혁신기획팀", "산업안전팀", "안전감사실", "경영지원팀","산장관광지","가평썰매장","칼봉산자연휴양림","연인산다목적캠핑장","자라섬캠핑장","교통약자이동지원센터","문화예술관","여성비전센터","종량제봉투판매","창업경제타운","한석봉체육관","생활체육파트","조종국민체육센터","청평호반문화체육센터","가평파크골프장"]
def extract_data_from_file(file_name):
    # 엑셀 파일 열기
    wb = load_workbook(filename=file_name)
    ws = wb.active

    # 데이터를 추출하여 리스트에 저장
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))  # 각 행의 데이터를 리스트로 변환하여 추가

    return data
def parse_week(week_string):
    # "2024년 1월 1주차" 같은 문자열을 year, month, week로 나눕니다.
    pattern = r"(\d{4})년 (\d{1,2})월 (\d)주차"
    match = re.match(pattern, week_string)
    
    if match:
        year = match.group(1)  # "2024"
        month = match.group(2)  # "1"
        week = match.group(3)  # "1"
        return year, month, week
    return None, None, None

def merge_all_with_keyword(worksheet, keyword='추진내용'):
    # '추진내용'을 포함한 모든 셀을 찾아 병합
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        for cell in row:
            if cell.value == keyword:
                start_row = cell.row
                start_col = cell.column

                # '추진내용' 셀과 그 오른쪽 셀 병합
                worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)

                # 병합된 셀에 '추진내용' 그대로의 값을 적용
                worksheet.cell(row=start_row, column=start_col, value=keyword)

                # 병합된 셀에 가운데 정렬과 테두리 적용
                merged_cell = worksheet.cell(row=start_row, column=start_col)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                merged_cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

def merge_excel_files(task_type, year, month, week):
    merged_wb = Workbook()
    merged_ws = merged_wb.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 부서별로 파일을 병합 (전주 업무 -> 금주 업무 순)
    files_in_dir = os.listdir()
    for department in departments:
        previous_task_type = "전주업무추진실적"
        previous_file_pattern = f"{previous_task_type}_{year}년_{month}월_{week}주차_{department}_report.xlsx"
        previous_files = [file for file in files_in_dir if previous_file_pattern in file]

        temp_previous_ws = None
        temp_current_ws = None

        # 전주 업무 파일 병합
        if previous_files:
            temp_previous_ws = Workbook().active
            for previous_file_name in previous_files:
                print(f"Merging file: {previous_file_name}")
                merge_file_into_workbook(temp_previous_ws, previous_file_name)

        # 금주 업무 파일 병합
        current_task_type = "금주업무추진계획"
        current_file_pattern = f"{current_task_type}_{year}년_{month}월_{week}주차_{department}_report.xlsx"
        current_files = [file for file in files_in_dir if current_file_pattern in file]

        if current_files:
            temp_current_ws = Workbook().active
            for current_file_name in current_files:
                print(f"Merging file: {current_file_name}")
                merge_file_into_workbook(temp_current_ws, current_file_name)

        # 데이터가 있는지 확인 후 헤더 추가
        if temp_previous_ws or temp_current_ws:
            # 전주와 금주 업무가 있을 때만 헤더를 추가
            last_row = merged_ws.max_row + 1
            merged_ws.append(["", "", "", ""])  # 첫 번째 빈 행 추가
            merged_ws.append(["", "", "", ""])  # 두 번째 빈 행 추가
            last_row = merged_ws.max_row
            merged_ws.merge_cells(f"A{last_row-1}:D{last_row}")
            merged_ws[f"A{last_row-1}"].value = f"<{department}>"
            merged_ws[f"A{last_row-1}"].alignment = Alignment(horizontal="center", vertical="center")
            merged_ws[f"A{last_row-1}"].font = Font(size=20, bold=True)

            # 전주 업무 데이터가 있을 때만 전주 업무 헤더 추가
            if temp_previous_ws:
                merged_ws.append([f"{year}년_{month}월_{week}주차 {department} 부서 전주 업무추진실적 "])
                last_row = merged_ws.max_row
                merged_ws.merge_cells(f"A{last_row}:D{last_row}")
                cell = merged_ws[f"A{last_row}"]
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border
                cell.font = Font(bold=True)

                last_row = merged_ws.max_row + 1
                merged_ws.append(["분류", "제목", "추진내용", "내용"])
                for col in range(1, 5):
                    cell = merged_ws.cell(row=last_row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                # 전주 업무 데이터 추가
                for row in temp_previous_ws.iter_rows(values_only=True):
                    merged_ws.append(row)

            # 금주 업무 데이터가 있을 때만 금주 업무 헤더 추가
            if temp_current_ws:
                last_row = merged_ws.max_row + 1
                merged_ws.append([f"{year}년_{month}월_{week}주차 {department} 부서 금주 업무추진계획 "])
                last_row = merged_ws.max_row
                merged_ws.merge_cells(f"A{last_row}:D{last_row}")
                cell = merged_ws[f"A{last_row}"]
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border
                cell.font = Font(bold=True)

                last_row = merged_ws.max_row + 1
                merged_ws.append(["분류", "제목", "추진내용", "내용"])
                for col in range(1, 5):
                    cell = merged_ws.cell(row=last_row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                # 금주 업무 데이터 추가
                for row in temp_current_ws.iter_rows(values_only=True):
                    merged_ws.append(row)

    

        start_row = None  # 병합을 시작할 행
    for row in range(2, merged_ws.max_row + 1):  # 2행부터 탐색 (헤더 제외)
        cell_value = merged_ws[f"A{row}"].value
        if cell_value is None and start_row is None:
            # 빈 셀을 처음 만나면 병합 시작을 설정
            start_row = row - 1
        elif cell_value is not None and start_row is not None:
            # 다음 데이터가 있을 경우 병합을 종료하고 A열과 B열 병합 수행
            # A열 병합
            merged_ws.merge_cells(f"A{start_row}:A{row-1}")
            for merge_row in range(start_row, row):
                merged_ws.cell(row=merge_row, column=1).border = thin_border
            
            # B열 병합 (A열과 동일한 범위로 병합)
            merged_ws.merge_cells(f"B{start_row}:B{row-1}")
            for merge_row in range(start_row, row):
                merged_ws.cell(row=merge_row, column=2).border = thin_border

            start_row = None  # 병합이 끝나면 초기화

    # 마지막 셀이 비어있으면 마지막까지 병합
    if start_row is not None:
        # C열과 D열에서 마지막 데이터가 있는 행 찾기
        last_data_row = max(
            merged_ws.max_row,
            max(row for row in range(2, merged_ws.max_row + 1)
                if merged_ws[f"C{row}"].value or merged_ws[f"D{row}"].value)
        )
        # A열 병합
        merged_ws.merge_cells(f"A{start_row}:A{last_data_row}")
        for merge_row in range(start_row, last_data_row + 1):
            merged_ws.cell(row=merge_row, column=1).border = thin_border
        
        # B열 병합 (A열과 동일한 범위로 병합)
        merged_ws.merge_cells(f"B{start_row}:B{last_data_row}")
        for merge_row in range(start_row, last_data_row + 1):
            merged_ws.cell(row=merge_row, column=2).border = thin_border

            # A열에 대해 다시 스캔하여 같은 값을 가진 행들을 병합하고 빈칸 처리


    

    merge_all_with_keyword(merged_ws, keyword='추진내용')

            
            


          
            


    column_widths = {
        'A': 10,
        'B': 12,
        'C': 10,
        'D': 55
    }

    for col, width in column_widths.items():
        merged_ws.column_dimensions[col].width = width

     # A, B, C열을 모두 가운데 정렬 적용 (병합되지 않은 셀에만 적용)
    for row in range(2, merged_ws.max_row + 1):  # 2행부터 시작 (헤더 제외)
        for col in range(1, 4):  # A(1), B(2), C(3) 열에 대해서
            cell = merged_ws.cell(row=row, column=col)
            if not merged_ws.merged_cells or cell.coordinate not in merged_ws.merged_cells:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for merge_cell in merged_ws.merged_cells.ranges:
     cell = merged_ws[merge_cell.coord.split(":")[0]]  # 병합 범위의 첫 번째 셀
    # 특정 부서 관련 텍스트가 포함된 셀은 제외
     if not (f"부서 금주 업무추진계획" in str(cell.value) or f"부서 전주 업무추진실적" in str(cell.value)):
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 모든 셀에 얇은 테두리 적용 (특정 텍스트가 포함된 셀은 제외)
    for row in merged_ws.iter_rows():
     for cell in row:
        # 특정 부서 관련 텍스트가 포함된 셀은 제외
        if not (f"부서 금주 업무추진계획" in str(cell.value) or f"부서 전주 업무추진실적" in str(cell.value)):
            cell.border = thin_border

    # 텍스트 줄바꿈 설정
    for row in merged_ws.iter_rows():
     for cell in row:
        if cell.alignment:
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                       vertical=cell.alignment.vertical,
                                       wrap_text=True)
    merged_ws.page_margins = PageMargins(
    left=1.1,    # 좌측 여백
    right=0.6,   # 우측 여백
    top=2.4,     # 위쪽 여백
    bottom=2.4   # 아래쪽 여백
)
        
    return merged_wb

def merge_file_into_workbook(merged_ws, file_name):
    """파일을 읽어서 데이터를 병합하는 함수"""
    wb = load_workbook(file_name)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        merged_ws.append(row)

   


@app.route('/download', methods=['POST'])
def download():
    download_week = request.form.get('download_week')
    task_type = request.form.get('taskType')

    # 주차 값 파싱
    year, month, week = parse_week(download_week)

    # 엑셀 파일 병합 처리
    merged_wb = merge_excel_files(task_type, year, month, week)

    # 파일 저장 경로 및 이름 설정
    output_file = f"static/{year}년_{month}월_{week}주차_주간업무보고.xlsx"

    # 병합된 파일을 저장
    try:
        if os.path.exists(output_file):
            os.remove(output_file)
        merged_wb.save(output_file)
        print(f"File saved at: {output_file}")
    except Exception as e:
        print(f"Error saving or deleting file: {e}")

    # 병합된 파일을 다운로드로 제공
    return send_file(output_file, as_attachment=True)




def merge_excel_filesv2(task_type, year, start_month, start_week, end_month, end_week, primarycategoryv2, secondarycategory, work_type2):
    merged_wb = Workbook()
    merged_ws = merged_wb.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 주차 범위 내의 모든 주차를 병합
    for month in range(int(start_month), int(end_month) + 1):
        for week in range(1, 6):  # 최대 5주차 가정
            if month == int(start_month) and week < int(start_week):
                continue
            if month == int(end_month) and week > int(end_week):
                break

            for department in departments:
                previous_task_type = "전주업무추진실적"
                previous_file_pattern = f"{previous_task_type}_{year}년_{month}월_{week}주차_{department}_report.xlsx"
                previous_files = [file for file in os.listdir() if previous_file_pattern in file]

                # 필터링된 파일만 처리
                filtered_previous_files = [file for file in previous_files if filter_files(primarycategoryv2, secondarycategory, work_type2, file)]

                temp_previous_ws = None
                if filtered_previous_files:
                    temp_previous_ws = Workbook().active
                    for previous_file_name in filtered_previous_files:
                        print(f"Merging file: {previous_file_name}")
                        merge_file_into_workbook(temp_previous_ws, previous_file_name)

                current_task_type = "금주업무추진계획"
                current_file_pattern = f"{current_task_type}_{year}년_{month}월_{week}주차_{department}_report.xlsx"
                current_files = [file for file in os.listdir() if current_file_pattern in file]

                # 필터링된 파일만 처리
                filtered_current_files = [file for file in current_files if filter_files(primarycategoryv2, secondarycategory, work_type2, file)]

                temp_current_ws = None
                if filtered_current_files:
                    temp_current_ws = Workbook().active
                    for current_file_name in filtered_current_files:
                        print(f"Merging file: {current_file_name}")
                        merge_file_into_workbook(temp_current_ws, current_file_name)

                # 헤더 추가 및 병합 처리 (전주 또는 금주 업무가 있을 경우에만 헤더 추가)
                if temp_previous_ws or temp_current_ws:
                    last_row = merged_ws.max_row + 1
                    merged_ws.append(["", "", "", ""])
                    merged_ws.append(["", "", "", ""])
                    last_row = merged_ws.max_row
                    merged_ws.merge_cells(f"A{last_row-1}:D{last_row}")
                    merged_ws[f"A{last_row-1}"].value = f"<{department}>"
                    merged_ws[f"A{last_row-1}"].alignment = Alignment(horizontal="center", vertical="center")
                    merged_ws[f"A{last_row-1}"].font = Font(size=20, bold=True)

                # 전주 업무추진실적 처리
                if temp_previous_ws:
                    merged_ws.append([f"{year}년_{month}월_{week}주차 {department} 부서 전주 업무추진실적 "])
                    last_row = merged_ws.max_row
                    merged_ws.merge_cells(f"A{last_row}:D{last_row}")
                    merged_ws[f"A{last_row}"].alignment = Alignment(horizontal="left", vertical="center")
                    merged_ws[f"A{last_row}"].font = Font(bold=True)

                    last_row = merged_ws.max_row + 1
                    merged_ws.append(["분류", "제목", "추진내용", "내용"])
                    for col in range(1, 5):
                        cell = merged_ws.cell(row=last_row, column=col)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border

                    for row in temp_previous_ws.iter_rows(values_only=True):
                        merged_ws.append(row)

                # 금주 업무추진계획 처리
                if temp_current_ws:
                    merged_ws.append([f"{year}년_{month}월_{week}주차 {department} 부서 금주 업무추진계획 "])
                    last_row = merged_ws.max_row
                    merged_ws.merge_cells(f"A{last_row}:D{last_row}")
                    merged_ws[f"A{last_row}"].alignment = Alignment(horizontal="left", vertical="center")
                    merged_ws[f"A{last_row}"].font = Font(bold=True)

                    last_row = merged_ws.max_row + 1
                    merged_ws.append(["분류", "제목", "추진내용", "내용"])
                    for col in range(1, 5):
                        cell = merged_ws.cell(row=last_row, column=col)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border

                    for row in temp_current_ws.iter_rows(values_only=True):
                        merged_ws.append(row)

    merge_all_with_keyword(merged_ws, keyword='추진내용')

    column_widths = {
        'A': 15,
        'B': 25,
        'C': 10,
        'D': 70
    }

    for col, width in column_widths.items():
        merged_ws.column_dimensions[col].width = width

    for row in range(2, merged_ws.max_row + 1):
        for col in range(1, 4):
            cell = merged_ws.cell(row=row, column=col)
            if not merged_ws.merged_cells or cell.coordinate not in merged_ws.merged_cells:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for merge_cell in merged_ws.merged_cells.ranges:
        cell = merged_ws[merge_cell.coord.split(":")[0]]
        if not (f"부서 금주 업무추진계획" in str(cell.value) or f"부서 전주 업무추진실적" in str(cell.value)):
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in merged_ws.iter_rows():
        for cell in row:
            if not (f"부서 금주 업무추진계획" in str(cell.value) or f"부서 전주 업무추진실적" in str(cell.value)):
                cell.border = thin_border
    start_row = None  # 병합을 시작할 행
    for row in range(2, merged_ws.max_row + 1):  # 2행부터 탐색 (헤더 제외)
        cell_value = merged_ws[f"A{row}"].value
        if cell_value is None and start_row is None:
            # 빈 셀을 처음 만나면 병합 시작을 설정
            start_row = row - 1
        elif cell_value is not None and start_row is not None:
            # 다음 데이터가 있을 경우 병합을 종료하고 A열과 B열 병합 수행
            # A열 병합
            merged_ws.merge_cells(f"A{start_row}:A{row-1}")
            for merge_row in range(start_row, row):
                merged_ws.cell(row=merge_row, column=1).border = thin_border
            
            # B열 병합 (A열과 동일한 범위로 병합)
            merged_ws.merge_cells(f"B{start_row}:B{row-1}")
            for merge_row in range(start_row, row):
                merged_ws.cell(row=merge_row, column=2).border = thin_border

            start_row = None  # 병합이 끝나면 초기화

    # 마지막 셀이 비어있으면 마지막까지 병합
    if start_row is not None:
        # C열과 D열에서 마지막 데이터가 있는 행 찾기
        last_data_row = max(
            merged_ws.max_row,
            max(row for row in range(2, merged_ws.max_row + 1)
                if merged_ws[f"C{row}"].value or merged_ws[f"D{row}"].value)
        )
        # A열 병합
        merged_ws.merge_cells(f"A{start_row}:A{last_data_row}")
        for merge_row in range(start_row, last_data_row + 1):
            merged_ws.cell(row=merge_row, column=1).border = thin_border
        
        # B열 병합 (A열과 동일한 범위로 병합)
        merged_ws.merge_cells(f"B{start_row}:B{last_data_row}")
        for merge_row in range(start_row, last_data_row + 1):
            merged_ws.cell(row=merge_row, column=2).border = thin_border
    for row in merged_ws.iter_rows():
        for cell in row:
            if cell.alignment:
                cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=True)

    return merged_wb



# primary와 second 배열 정의
primary = {"리더쉽전략", "경영시스템", "사회적책임", "주요사업성과", "경영효율성과", "고객만족성과", "권장정책성과"}
second = {"경영층의리더쉽", "전략및혁신", "조직인사관리", "윤리경영", "재무관리", "지역상생협력", "안전및환경", "소통및참여", "주요사업성과", "경영효율성과", "고객만족도", "권장정책목적달성도"}

def extract_work_type_and_categories(file_name):
    # work_type2는 '주요업무' 또는 '일반업무'와 같은 값을 포함한다고 가정
    work_type_pattern = r'(주요업무|일반업무)'  # 추출할 work_type2 패턴

    # work_type2 추출
    work_type_match = re.search(work_type_pattern, file_name)
    work_type_value = work_type_match.group(0) if work_type_match else None

    # 검색할 패턴을 결정하는 로직
 #   if firstcategory == "전체":
        # firstcategory가 '전체'인 경우: primary-second로 검색
    category_pattern = r'({})-({})'.format('|'.join(primary), '|'.join(second))
 #   elif secondcategory == "전체":
        # firstcategory가 '전체'가 아니고 secondcategory가 '전체'인 경우: firstcategory-second로 검색
  #      category_pattern = r'({})-({})'.format(firstcategory, '|'.join(second))
 #   else:
        # firstcategory와 secondcategory 모두 '전체'가 아닌 경우: firstcategory-secondcategory로 검색
  #      category_pattern = r'({})-({})'.format(firstcategory, secondcategory)

    # {firstcategory}-{secondcategory} 패턴을 모두 추출
    category_matches = re.findall(category_pattern, file_name)
    categories = [f"{first_}-{second_}" for first_, second_ in category_matches]
    print(f"카테고리 : {category_pattern}")

    evaluation_keywords = ['Plan', 'Do', 'Check', 'Act']
    evaluation_process_matches = [keyword for keyword in evaluation_keywords if keyword in file_name]
    evaluation_process = ' / '.join(evaluation_process_matches) if evaluation_process_matches else None

    relationship_keywords = ['자치단체 의회', '지역유관기관', '내부구성원', '지역주민','일반국민']
    relationship_matches = [keyword for keyword in relationship_keywords if keyword in file_name]
    relationship = ' / '.join(relationship_matches) if relationship_matches else None


   

    return work_type_value, categories, evaluation_process, relationship


def filter_files(primarycategoryv2, secondarycategory, work_type2, file_name):
    # 1차 필터링: WorkType 확인 ('전체', '주요업무', '일반업무')
    if work_type2 != "전체":
        if work_type2 == "주요업무" and "주요업무" not in file_name:
            return False  # '주요업무'가 포함되지 않은 파일 제외
        elif work_type2 == "일반업무" and "일반업무" not in file_name:
            return False  # '일반업무'가 포함되지 않은 파일 제외

    # 2차 필터링: primarycategoryv2와 secondcategory 확인
    if primarycategoryv2 != "전체":
        if primarycategoryv2 not in file_name:
            return False  # primarycategoryv2가 포함되지 않은 파일 제외
        if secondarycategory != "전체" and secondarycategory not in file_name:
            return False  # secondcategory가 포함되지 않은 파일 제외

    return True  # 모든 조건을 통과한 파일만 반환



def merge_excel_filesv3(task_type, year, start_month, start_week, end_month, end_week):
    merged_wb = Workbook()
    merged_ws = merged_wb.active

    # 헤더 정의 ('업무 형태'와 '업무 카테고리' 추가됨)
    headers = ['담당', '제목', '대상', '대상유형', '기간', '장소', '기대효과/추진성과', '내용', '목적', '행정사항', '소요예상', '관련자료', '업무 형태', '업무 카테고리', '경영평가상 절차']
    
    # 병합된 시트의 첫 번째 행에 헤더 추가
    merged_ws.append(headers)
    # 헤더에 굵은 글씨체와 가운데 정렬 서식 적용
    for col_num, header_cell in enumerate(merged_ws[1], 1):
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 주차 범위 내의 모든 주차를 병합
    for month in range(int(start_month), int(end_month) + 1):
        for week in range(1, 6):  # 최대 5주차 가정
            if month == int(start_month) and week < int(start_week):
                continue
            if month == int(end_month) and week > int(end_week):
                break

            for department in departments:
                previous_task_type = "전주업무추진실적"
                previous_file_pattern = f"{previous_task_type}_{year}년_{month}월_{week}주차_{department}_report.xlsx"
                previous_files = list(set([file for file in os.listdir() if previous_file_pattern in file]))

                # 전주 업무추진실적 처리
                for previous_file_name in previous_files:
                    # 실제 파일을 읽고 데이터를 병합
                    temp_previous_ws = load_workbook(previous_file_name).active
                    print(f"Merging file: {previous_file_name}")

                    # 파일명에서 '업무 형태'와 '업무 카테고리' 추출
                    work_type_value, category_values, evaluation_process, relationship = extract_work_type_and_categories(previous_file_name)
                    category_value = '/'.join(category_values) if category_values else ''

                    # 새로운 행 생성
                    current_row = {header: '' for header in headers}
                    current_row['업무 형태'] = work_type_value
                    current_row['업무 카테고리'] = category_value
                    current_row['경영평가상 절차'] = evaluation_process
                    current_row['대상유형'] = relationship

                    # 워크시트의 행을 반복하며 데이터를 current_row에 추가
                    for row in temp_previous_ws.iter_rows(values_only=True):
                        # 완전히 빈 행인 경우 새로운 데이터 블록 시작
                        if not any(row):
                            if any(current_row.values()):  # 현재 데이터가 비어 있지 않다면 저장
                                merged_ws.append(list(current_row.values()))
                                # 새로운 데이터 블록을 위한 행 초기화 및 공통 값 재적용
                                current_row = {header: '' for header in headers}
                                current_row['업무 형태'] = work_type_value
                                current_row['업무 카테고리'] = category_value
                                current_row['경영평가상 절차'] = evaluation_process
                                current_row['대상유형'] = relationship
                            continue

                        # 새로운 데이터 블록의 시작 여부를 확인 ('담당' 값이 새로 나타난 경우)
                        if row[0] and current_row['담당']:
                            merged_ws.append(list(current_row.values()))
                            current_row = {header: '' for header in headers}
                            current_row['업무 형태'] = work_type_value
                            current_row['업무 카테고리'] = category_value
                            current_row['경영평가상 절차'] = evaluation_process
                            current_row['대상유형'] = relationship

                        # 워크시트에서 값을 매핑하여 current_row에 저장
                        if row[0]:
                            current_row['담당'] = row[0]
                        if row[1]:
                            current_row['제목'] = row[1]
                        if len(row) > 3 and row[2] in headers and row[3] is not None:
                            current_row[row[2]] = row[3]
                    
                    # 마지막으로 처리된 current_row를 병합된 데이터에 추가
                    if any(current_row.values()):  # 마지막 데이터가 비어 있지 않다면 추가
                        merged_ws.append(list(current_row.values()))


    return merged_wb


    

@app.route('/downloadreport1', methods=['POST'])
def downloadreport1():
    download_start_week = request.form.get('start_week')
    download_end_week = request.form.get('end_week')
    task_type = request.form.get('taskType')
    work_type2 = request.form.get('work_type2')
    #primarycategory = request.form.get('primary_category')
    #secondarycategory = request.form.get('secondary_category')
    #primarycategoryv2 = primarycategory.replace('.', '')
   
    #print(f"{secondarycategory}")

    start_year, start_month, start_week = parse_week(download_start_week)
    end_year, end_month, end_week = parse_week(download_end_week)

    merged_wb = merge_excel_filesv3(task_type, start_year, start_month, start_week, end_month,end_week)
    output_file = f"static/{start_year}년_{start_month}월_{start_week}주차_~_{end_year}년_{end_month}월_{end_week}주차_주간업무보고.xlsx"
    try:
        if os.path.exists(output_file):
            os.remove(output_file)
        merged_wb.save(output_file)
        print(f"File saved at: {output_file}")
    except Exception as e:
        print(f"Error saving or deleting file: {e}")

    return send_file(output_file, as_attachment=True)



def merge_excel_files_for_month(year, month):
    merged_wb = Workbook()
    merged_ws = merged_wb.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    files_in_dir = os.listdir()

    for department in departments:
        department_header_added = False
        for week in range(1, 6):  # 최대 5주차까지 검색
            previous_task_type = "전주업무추진실적"
            previous_file_pattern = f"{previous_task_type}_{year}년_{month}월_{week}주차_{department}_report.xlsx"
            previous_files = [file for file in files_in_dir if previous_file_pattern in file]

           
            temp_previous_ws = None
           

            # 전주 업무 병합
            if previous_files:
                temp_previous_ws = Workbook().active
                for previous_file_name in previous_files:
                    print(f"Merging file: {previous_file_name}")
                    merge_file_into_workbook(temp_previous_ws, previous_file_name)

           

            # 데이터가 있는지 확인 후 헤더 추가
            if temp_previous_ws:
                if not department_header_added:
                 last_row = merged_ws.max_row + 1
                 merged_ws.append(["", "", "", ""])  # 빈 행 추가
                 merged_ws.append(["", "", "", ""])  # 빈 행 추가
                 last_row = merged_ws.max_row
                 merged_ws.merge_cells(f"A{last_row-1}:D{last_row}")
                 merged_ws[f"A{last_row-1}"].value = f"<{department}>"
                 merged_ws[f"A{last_row-1}"].alignment = Alignment(horizontal="center", vertical="center")
                 merged_ws[f"A{last_row-1}"].font = Font(size=20, bold=True)
                 department_header_added = True  # 부서 헤더가 추가된 것으로 플래그 설정

               
                merged_ws.append([f"{year}년_{month}월_{week}주차 {department} 부서 업무추진실적 "])
                last_row = merged_ws.max_row
                merged_ws.merge_cells(f"A{last_row}:D{last_row}")
                cell = merged_ws[f"A{last_row}"]
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(bold=True)

                last_row = merged_ws.max_row + 1
                merged_ws.append(["분류", "제목", "추진내용", "내용"])
                for col in range(1, 5):
                    cell = merged_ws.cell(row=last_row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                for row in temp_previous_ws.iter_rows(values_only=True):
                        merged_ws.append(row)


                start_row = None  # 병합을 시작할 행
                # 데이터를 병합하는 부분
                for row in range(2, merged_ws.max_row + 1):  # 2행부터 탐색 (헤더 제외)
                    cell_value = merged_ws[f"A{row}"].value
                    if cell_value is None and start_row is None:
                        start_row = row - 1
                    elif cell_value is not None and start_row is not None:
                        merged_ws.merge_cells(f"A{start_row}:A{row-1}")
                        for merge_row in range(start_row, row):
                            merged_ws.cell(row=merge_row, column=1).border = thin_border
                        merged_ws.merge_cells(f"B{start_row}:B{row-1}")
                        for merge_row in range(start_row, row):
                            merged_ws.cell(row=merge_row, column=2).border = thin_border
                        start_row = None

                if start_row is not None:
                    last_data_row = max(
                        merged_ws.max_row,
                        max(row for row in range(2, merged_ws.max_row + 1)
                            if merged_ws[f"C{row}"].value or merged_ws[f"D{row}"].value)
                    )
                    merged_ws.merge_cells(f"A{start_row}:A{last_data_row}")
                    for merge_row in range(start_row, last_data_row + 1):
                        merged_ws.cell(row=merge_row, column=1).border = thin_border
                    merged_ws.merge_cells(f"B{start_row}:B{last_data_row}")
                    for merge_row in range(start_row, last_data_row + 1):
                        merged_ws.cell(row=merge_row, column=2).border = thin_border

                merge_all_with_keyword(merged_ws, keyword='추진내용')
        column_widths = {
        'A': 15,
        'B': 25,
        'C': 10,
        'D': 60
    }

    for col, width in column_widths.items():
        merged_ws.column_dimensions[col].width = width

     # A, B, C열을 모두 가운데 정렬 적용 (병합되지 않은 셀에만 적용)
    for row in range(2, merged_ws.max_row + 1):  # 2행부터 시작 (헤더 제외)
        for col in range(1, 4):  # A(1), B(2), C(3) 열에 대해서
            cell = merged_ws.cell(row=row, column=col)
            if not merged_ws.merged_cells or cell.coordinate not in merged_ws.merged_cells:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for merge_cell in merged_ws.merged_cells.ranges:
     cell = merged_ws[merge_cell.coord.split(":")[0]]  # 병합 범위의 첫 번째 셀
    # 특정 부서 관련 텍스트가 포함된 셀은 제외
     if not (f"부서 금주 업무추진계획" in str(cell.value) or f"부서 업무추진실적" in str(cell.value)):
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 모든 셀에 얇은 테두리 적용 (특정 텍스트가 포함된 셀은 제외)
    for row in merged_ws.iter_rows():
     for cell in row:
        # 특정 부서 관련 텍스트가 포함된 셀은 제외
        if not (f"부서 금주 업무추진계획" in str(cell.value) or f"부서 업무추진실적" in str(cell.value)):
            cell.border = thin_border

    # 텍스트 줄바꿈 설정
    for row in merged_ws.iter_rows():
     for cell in row:
        if cell.alignment:
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                       vertical=cell.alignment.vertical,
                                       wrap_text=True)
        
    return merged_wb




def parse_month(week_string):
    # 문자열을 다듬고, 불필요한 공백 등을 제거
    week_string = week_string.strip()

    # 정규 표현식을 사용해 연도와 월을 추출 (더 유연하게 처리)
    pattern = r"(\d{4})년\s*(\d{1,2})월"
    match = re.match(pattern, week_string)
    
    if match:
        year = match.group(1)  # "2024"
        month = match.group(2)  # "1"
        return year, month
    return None, None



@app.route('/download2', methods=['POST'])
def download2():
    months = request.form.get('month')
    year, month = parse_month(months)   

    if not year or not month:
        return "Invalid month format.", 400

    merged_wb = merge_excel_files_for_month(year, month)

    output_file = f"static/{year}년_{month}월_월간업무보고.xlsx"
    
    try:
        if os.path.exists(output_file):
            os.remove(output_file)
        merged_wb.save(output_file)
        print(f"File saved at: {output_file}")
    except Exception as e:
        print(f"Error saving or deleting file: {e}")
        return str(e), 500

    return send_file(output_file, as_attachment=True)

@app.route('/downloadpdf/details')
def download_details():
    path = r"C:\Users\user\Desktop\Project1\static\img\설명서.pdf"  # 다운로드할 파일 경로
    return send_file(path, as_attachment=True)

if __name__ == '__main__':
  app.run(host='0.0.0.0', port=5000, debug=False)
